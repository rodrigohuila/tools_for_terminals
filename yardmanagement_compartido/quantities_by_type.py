#! python3

import os
from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import *  # Prueba borrar cuadro dialogo
from tkinter.ttk import * # Prueba borrar cuadro dialogo
from os.path import basename
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd
import xlwt
from xlutils.copy import copy
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import pyexcel as p
from xls2xlsx import XLS2XLSX
from operator import itemgetter, attrgetter
from datetime import datetime, date, timedelta
from functools import reduce
import numpy as np
import pandas as pd
import csv


#pathFile = os.getcwd()
#print(pathFile)
#root = Tk() # ERASE DIALOG
#root.update()

def clear():  # ERASE SCREEN
    if os.name == "nt":      # Windows
        os.system("cls")
    else:
        os.system("clear")   # Linux


def openFile():  # DIALOG TO CHOOSE A FILE
    excel_file = FileDialog.askopenfilename(
        initialdir=pathFile,
        filetypes=(("xls", "*.xls"), ("all files", "*.*")),
        title="Escoja el archivo de Existencias Excel a Procesar"
    )
    base = os.path.basename(excel_file)
    name = os.path.splitext(base)[0]
    print("\nEl nombre del archivo inicial es: " + name + "\n")
    return(excel_file)


def count_quantity(pies): # figure out quantity
     return 1           



def quantities_list(sheet, pathtoSave): # RECEIVE EXCEL SHIFT AND GENERETE A ORDERED LIST
    
    #INIT VAR AND LIST
    containers_import = []
    containers_export = []
    containers_mty = []
    containers_trb = []
    containers_reefer = []
    containers_ovh = []
    containers_ovhemt = []
    containers_apto = []
    containers_all = []
    containers_lost = []
    total_cantidades = []

    total_vacios = 0
    total_llenos = 0
    
    now = datetime.now()
    formato = now.strftime("%d-%m-%Y %H-%M")
           
    # Loop through the sheet
    for i in range(sheet.nrows):
        frigo = (repr(sheet.cell_value(i, 18)).replace("'",""))
        sit = (repr(sheet.cell_value(i, 8)).replace("'",""))
        pdescarga = (repr(sheet.cell_value(i, 6)).replace("'",""))
        pfinal = (repr(sheet.cell_value(i, 7)).replace("'",""))
        tipo = (repr(sheet.cell_value(i, 2)).replace("'",""))
        iso = (repr(sheet.cell_value(i, 1)).replace("'",""))
        pies = (repr(sheet.cell_value(i, 3)))
        estatus = (repr(sheet.cell_value(i, 9)).replace("'",""))
        retencion = (repr(sheet.cell_value(i, 19)).replace("'",""))
        ubicacion = (repr(sheet.cell_value(i, 21)).replace("'",""))
        zona = (ubicacion[0:1])
        bloque = (ubicacion[2:4])
        modulo = (ubicacion[5:8])
        calle = (ubicacion[9:12])
        altura = (ubicacion[12:13])
        linea = (repr(sheet.cell_value(i, 10)).replace("'",""))
        dias_terminal = repr(sheet.cell_value(i, 24)).replace("'","")
        container = repr(sheet.cell_value(i, 0))
        buque_import = (repr(sheet.cell_value(i, 11)).replace("'",""))
        viaje_import = (repr(sheet.cell_value(i, 12)).replace("'",""))
        buque_export = (repr(sheet.cell_value(i, 13)).replace("'",""))
        viaje_export = (repr(sheet.cell_value(i, 14)).replace("'",""))


        if sit == 'C' or sit == '' or sit == 'E':
           pies = int(float(pies))
           dias_terminal = int(float(dias_terminal))
           fecha = ((now - timedelta(days=dias_terminal)).strftime("%d-%m-%Y"))
           line = {"contenedor": container, "pies": pies, "tipo": tipo, "estatus": estatus,"pdescarga": pdescarga, "pfinal":  pfinal,
                       "sit": sit, "linea":linea, "buque_export" : buque_export, "viaje_export" : viaje_export , "fecha_ingreso": fecha}
           total_cantidades.append(line)       
           if ubicacion[0:4] != "S 04":
                if frigo == "":
                    if tipo != "FLT" and tipo != "HFL" and tipo != "O/T" and tipo != "OTH" and tipo != "T/K" and tipo != "":                       
                            if estatus == "HH": # Import
                                if pdescarga == "COBUN":                                   
                                        quantity = count_quantity(pies)
                                        line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea}
                                        containers_import.append(line)
                                        containers_all.append(line)
                                        total_llenos += int(quantity)
                                elif pdescarga != "COBUN": #Exportación                                    
                                            quantity = count_quantity(pies)
                                            line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea, "pdescarga": pdescarga}
                                            containers_export.append(line)
                                            containers_all.append(line)
                                            total_llenos += int(quantity)
                            elif estatus == "EMT" or estatus == "TRV":  # Empty
                                if pdescarga != "COCAF" and pdescarga != "COVAR" and pdescarga != "AZUCA" and pdescarga != "COSUG":                                    
                                        quantity = count_quantity(pies)
                                        line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea, "pdescarga": pdescarga}
                                        containers_mty.append(line)
                                        containers_all.append(line)
                                        total_vacios += int(quantity)
                                elif pdescarga == "COCAF" or pdescarga == "COVAR" or pdescarga == "AZUCA" or pdescarga == "COSUG": # Aptos                                   
                                        quantity = count_quantity(pies)
                                        line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies, "tipo": tipo, "estatus": estatus, "linea": linea, "pdescarga": pdescarga}
                                        containers_apto.append(line)
                                        containers_all.append(line)
                                        total_vacios += int(quantity)
                            elif estatus == "TRB": # Transbordos                                
                                    quantity = count_quantity(pies)
                                    line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea, "pdescarga": pdescarga}
                                    containers_trb.append(line)
                                    containers_all.append(line)
                                    total_llenos += int(quantity)
                    else:
                        if estatus == "HH" or estatus == "TRB":
                            quantity = count_quantity(pies) # OVH
                            line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea}
                            containers_ovh.append(line)
                            containers_all.append(line)
                            total_llenos += int(quantity)
                            #print (line)
                        else:
                            quantity = count_quantity(pies) # OVH EMTY
                            line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea}
                            containers_ovhemt.append(line)
                            containers_all.append(line)
                            total_vacios += int(quantity)
                            #print (line)
                elif frigo != "": # Refrigerados                    
                         quantity = count_quantity(pies)
                         line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea, "frigo": frigo}
                         containers_reefer.append(line)
                         containers_all.append(line)
                         total_llenos += int(quantity)
           else:
                quantity = count_quantity(pies)
                line = {"zona": zona, "bloque" : bloque,  "quantity": quantity, "pies": pies,"tipo": tipo, "estatus": estatus, "linea": linea}
                containers_lost.append(line)
                
    print ('Unidades perdidas:')
    for container in containers_lost:
        print (container)

    #os.chdir(pathFile) # go to Pathhfile

    #allContainers = {**containers_import, **containers_mty, **containers_apto, **containers_trb, **containers_export,  **containers_reefer, **containers_ovh, **containers_ovhemt}
                     
    df2import = pd.DataFrame(containers_import)
    df2empty = pd.DataFrame(containers_mty)
    df2apto = pd.DataFrame(containers_apto)
    df2trb = pd.DataFrame(containers_trb)
    df2export = pd.DataFrame(containers_export)
    df2reefer = pd.DataFrame(containers_reefer)
    df2ovh = pd.DataFrame(containers_ovh)
    df2ovhemt = pd.DataFrame(containers_ovhemt)
    df2lineasall = pd.DataFrame(containers_all)
    df = df2import + df2empty + df2apto + df2trb +df2export + df2reefer +  df2ovh + df2ovhemt + df2lineasall
    
    if len(containers_export) > 0:
        print ('\nEXPORTACIÓN')
        print (pd.pivot_table(df2export, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_trb) > 0:    
        print ('\nTRANSBORDOS')
        print (pd.pivot_table(df2trb, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_import) > 0:     
        print ('\nIMPORTACIÓN')
        print (pd.pivot_table(df2import, values='quantity', index=['estatus', 'pies', 'tipo'], aggfunc=np.sum, margins=True))
        ornament()
    if len(containers_reefer) > 0:     
        print ('\nREFRIGERADOS')
        print (pd.pivot_table(df2reefer, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()  
    if len(containers_ovh) > 0: 
        print ('\nOVH')
        print (pd.pivot_table(df2ovh, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_ovhemt) > 0: 
        print ('\nOVHEMT')
        print (pd.pivot_table(df2ovhemt, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_mty) > 0:     
        print ('\nVACÍOS')
        print (pd.pivot_table(df2empty, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_apto) > 0:     
        print ('\nAPTOS')
        print (pd.pivot_table(df2apto, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_all) > 0:  
        print ('\nCANTIDADDES POR TIPO: ')
        print (pd.pivot_table(df2lineasall, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    
    # Generar archivo EXCEL
    os.chdir(pathtoSave) #Go to the pathfile
    writer = pd.ExcelWriter(f'ocupación_por_tipo_{formato}.xlsx', engine='xlsxwriter')
    workbook = writer.book
    #worksheet = writer.sheets['Sheet11']
    writer.save()

    with pd.ExcelWriter(f'ocupación_por_tipo_{formato}.xlsx', engine='openpyxl', mode='a') as writer:
        
        (pd.pivot_table(df2export, values='quantity', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=1, startrow=1)
        (pd.pivot_table(df2trb, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=7, startrow=1)  
        (pd.pivot_table(df2import, values='quantity', index=['estatus', 'pies', 'tipo'], aggfunc=np.sum, margins=True)).to_excel(writer, header=True, startcol=1, startrow=18) 
        (pd.pivot_table(df2reefer, values='quantity', index=['estatus', 'pies', 'tipo', 'frigo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=7, startrow=18) 
        (pd.pivot_table(df2ovh, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=1, startrow=29)  
        (pd.pivot_table(df2ovhemt, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=7, startrow=29)    
        (pd.pivot_table(df2empty, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=14, startrow=1)   
        (pd.pivot_table(df2apto, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=14, startrow=29)   
        (pd.pivot_table(df2lineasall, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=21, startrow=1)


    book = openpyxl.load_workbook(f'ocupación_por_tipo_{formato}.xlsx')
    sheet = book.worksheets[1]
    ft = Font(name='Verdana',
                size=12,
                bold = True,
                color='FF000000')
    sheet['B1'] =  'EXPORTACIÓN'
    sheet['B1'].font = ft
    sheet['H1'] =  'TRANSBORDO'
    sheet['H1'].font = ft
    sheet['B18'] =  'IMPORTACIÓN'
    sheet['B18'].font = ft
    sheet['H18'] =  'REFRIGERADOS'
    sheet['H18'].font = ft
    sheet['B29'] =  'OVER HIGH FULL'
    sheet['B29'].font = ft
    sheet['H29'] =  'OVER HIGH EMTY'
    sheet['H29'].font = ft
    sheet['O1'] =  'VACÍOS'
    sheet['O1'].font = ft
    sheet['O29'] =  'APTO ALIMENTO'
    sheet['O29'].font = ft
    sheet['V1'] =  'ALL'
    sheet['V1'].font = ft
    sheet.title = 'Cantidades'

    sheet0 = book.worksheets[0]
    book.remove_sheet(sheet0)
     
    book.save(f'ocupación_por_tipo_{formato}.xlsx')



    # Generar archivo CSV
    #os.chdir(pathtoSave) #Go to the pathfile
    #my_file = f'total_cantidades_{formato}.csv'
    
    #(pd.pivot_table(df2export, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')    
    #(pd.pivot_table(df2trb, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')    
    #(pd.pivot_table(df2import, values='quantity', index=['estatus', 'pies', 'tipo'], aggfunc=np.sum, margins=True)).to_csv(my_file, mode='a')   
    #(pd.pivot_table(df2reefer, values='quantity', index=['estatus', 'pies', 'tipo', 'frigo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')    
    #(pd.pivot_table(df2ovh, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')   
    #(pd.pivot_table(df2ovhemt, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')    
    #(pd.pivot_table(df2empty, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')    
    #(pd.pivot_table(df2apto, values='quantity', index=['estatus', 'pdescarga', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')    
    #(pd.pivot_table(df2lineasall, values='quantity', index=['estatus', 'pies', 'tipo'], margins=True, aggfunc=np.sum)).to_csv(my_file, mode='a')       
         
       
    
    #TOTAL quantity
    llenos = int(total_llenos)
    vacios = int(total_vacios)
    total = llenos + vacios
    
    print (f"\nTotal Llenos {llenos}")
    print (f"Total Vacíos {vacios}")
    print (f"\nTotal Ocupación es {total} contenedores")

    return (vacios, llenos)            

   
def ornament():
    print("_" * 50)


def menu1():
    ornament()
    print("OCUPACIÓN EN quantity")
    ornament()
      

  
def main():  # MAIN
    clear()  
    menu1()
    excel_file= openFile()  # Name of the file to choose
    root.destroy() # Close Dialog
    wb = xlrd.open_workbook(excel_file, encoding_override="cp1252") # Read excel book
    sheet = wb.sheet_by_index(0)
    #ornament()  
    #ocupacion = ocupacion_list(sheet)
    #cap_llenos = 7150
    #cap_vacios = 6460
    #cap_total = 13610
    #llenos = int(ocupacion[1])
    #pllenos = (ocupacion[1] / cap_llenos) * 100
    #vacios = int(ocupacion[0])
    #pvacios = (ocupacion[0] / cap_vacios) * 100
    #total = llenos + vacios
    #ocupacion = (sum(ocupacion) / cap_total ) * 100
    #print (f"\nTotal Llenos {llenos} - ocupacion {round(pllenos, 1)}%")
    #print (f"Total Vacíos {vacios} - ocupacion {round(pvacios, 1)}%")
    #print (f"\nTotal Ocupación es {total} - {round(ocupacion, 1)}%")


if __name__ == '__main__':
    main()


#os.system("pause") # Press a key to continue 
