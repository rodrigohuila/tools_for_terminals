﻿#! python3

import os
from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import *  # Prueba borrar cuadro dialogo
from tkinter.ttk import * # Prueba borrar cuadro dialogo
from os.path import basename
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import xlrd
import xlwt
from xlutils.copy import copy
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import pyexcel as p
from xls2xlsx import XLS2XLSX
from operator import itemgetter, attrgetter
from datetime import datetime, date, timedelta
from functools import reduce
import numpy as np
import pandas as pd
import csv


#pathFile = os.getcwd()
#print(pathFile)
#root = Tk() # ERASE DIALOG
#root.update()

def clear():  # ERASE SCREEN
    if os.name == "nt":      # Windows
        os.system("cls")
    else:
        os.system("clear")   # Linux


def openFile():  # DIALOG TO CHOOSE A FILE
    excel_file = FileDialog.askopenfilename(
        initialdir=pathFile,
        filetypes=(("xls", "*.xls"), ("all files", "*.*")),
        title="Escoja el archivo de Existencias Excel a Procesar"
    )
    base = os.path.basename(excel_file)
    name = os.path.splitext(base)[0]
    print("\nEl nombre del archivo inicial es: " + name + "\n")
    return(excel_file)


def count_teus(pies): # figure out Teus
     #print (int(float(pies)))
     if int(float(pies)) == 20:
         return 1
     else:
         return 2         



def ocupacion_list(sheet, pathtoSave): # RECEIVE EXCEL SHIFT AND GENERETE A ORDERED LIST
    
    #INIT VAR AND LIST
    containers_import = []
    containers_export = []
    containers_mty = []
    containers_trb = []
    containers_reefer = []
    containers_ovh = []
    containers_ovhemt = []
    containers_apto = []
    containers_all = []
    containers_lost = []

    total_vacios = 0
    total_llenos = 0
    
    now = datetime.now()
    formato = now.strftime("%d-%m-%Y %H-%M")
           
    # Loop through the sheet
    for i in range(sheet.nrows):
        frigo = (repr(sheet.cell_value(i, 18)).replace("'",""))
        sit = (repr(sheet.cell_value(i, 8)).replace("'",""))
        pdescarga = (repr(sheet.cell_value(i, 6)).replace("'",""))
        tipo = (repr(sheet.cell_value(i, 2)).replace("'",""))
        iso = (repr(sheet.cell_value(i, 1)).replace("'",""))
        pies = (repr(sheet.cell_value(i, 3)))
        estatus = (repr(sheet.cell_value(i, 9)).replace("'",""))
        retencion = (repr(sheet.cell_value(i, 19)).replace("'",""))
        ubicacion = (repr(sheet.cell_value(i, 21)).replace("'",""))
        zona = (ubicacion[0:1])
        bloque = (ubicacion[2:4])
        modulo = (ubicacion[5:8])
        calle = (ubicacion[9:12])
        altura = (ubicacion[12:13])
        linea = (repr(sheet.cell_value(i, 10)).replace("'",""))
        dias_terminal = repr(sheet.cell_value(i, 24)).replace("'","")
        container = repr(sheet.cell_value(i, 0))

        if sit == "C" or sit == "":
            #if zona != "S" :
            if ubicacion[0:4] != "S 04":
                if frigo == "":
                    if tipo != "FLT" and tipo != "HFL" and tipo != "O/T" and tipo != "OTH" and tipo != "T/K" and tipo != "":                       
                            if estatus == "HH": # Import
                                if pdescarga == "COBUN":                                   
                                        teus = count_teus(pies)
                                        line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "linea": linea}
                                        containers_import.append(line)
                                        containers_all.append(line)
                                        total_llenos += int(teus)
                                elif pdescarga != "COBUN": #Exportación                                    
                                            teus = count_teus(pies)
                                            line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "linea": linea}
                                            containers_export.append(line)
                                            containers_all.append(line)
                                            total_llenos += int(teus)
                            elif estatus == "EMT" or estatus == "TRV":  # Empty
                                if pdescarga != "COCAF" and pdescarga != "COVAR" and pdescarga != "AZUCA" and pdescarga != "COSUG":                                    
                                        teus = count_teus(pies)
                                        line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "descarga": pdescarga, "linea": linea}
                                        containers_mty.append(line)
                                        containers_all.append(line)
                                        total_vacios += int(teus)
                                elif pdescarga == "COCAF" or pdescarga == "COVAR" or pdescarga == "AZUCA" or pdescarga == "COSUG": # Aptos                                   
                                        teus = count_teus(pies)
                                        line = {"zona": zona, "bloque": bloque,  "teus": teus, "estatus": estatus, "descarga": pdescarga, "linea": linea}
                                        containers_apto.append(line)
                                        containers_all.append(line)
                                        total_vacios += int(teus)
                            elif estatus == "TRB": # Transbordos                                
                                    teus = count_teus(pies)
                                    line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "linea": linea}
                                    containers_trb.append(line)
                                    containers_all.append(line)
                                    total_llenos += int(teus)
                    else:
                        if estatus == "HH" or estatus == "TRB":
                            teus = count_teus(pies) # OVH
                            line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "tipo": tipo, "linea": linea}
                            containers_ovh.append(line)
                            containers_all.append(line)
                            total_llenos += int(teus)
                            #print (line)
                        else:
                            teus = count_teus(pies) # OVH EMTY
                            line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "tipo": tipo, "linea": linea}
                            containers_ovhemt.append(line)
                            containers_all.append(line)
                            total_vacios += int(teus)
                            #print (line)
                elif frigo != "": # Refrigerados                    
                         teus = count_teus(pies)
                         line = {"zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "frigo": frigo, "linea": linea}
                         containers_reefer.append(line)
                         containers_all.append(line)
                         total_llenos += int(teus)
            else:
                teus = count_teus(pies)
                line = {"container": container, "zona": zona, "bloque" : bloque,  "teus": teus, "estatus": estatus, "linea": linea}
                containers_lost.append(line)
                
    print ('Unidades perdidas:')
    for container in containers_lost:
        print (container)

    #os.chdir(pathFile) # go to Pathhfile

    #allContainers = {**containers_import, **containers_mty, **containers_apto, **containers_trb, **containers_export,  **containers_reefer, **containers_ovh, **containers_ovhemt}
                     
    df2import = pd.DataFrame(containers_import)
    df2empty = pd.DataFrame(containers_mty)
    df2apto = pd.DataFrame(containers_apto)
    df2trb = pd.DataFrame(containers_trb)
    df2export = pd.DataFrame(containers_export)
    df2reefer = pd.DataFrame(containers_reefer)
    df2ovh = pd.DataFrame(containers_ovh)
    df2ovhemt = pd.DataFrame(containers_ovhemt)
    df2lineasall = pd.DataFrame(containers_all)
    df = df2import + df2empty + df2apto + df2trb +df2export + df2reefer +  df2ovh + df2ovhemt + df2lineasall
    
    if len(containers_export) > 0:
        print ('\nEXPORTACIÓN')
        print (pd.pivot_table(df2export, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_trb) > 0:    
        print ('\nTRANSBORDOS')
        print (pd.pivot_table(df2trb, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_import) > 0:     
        print ('\nIMPORTACIÓN')
        print (pd.pivot_table(df2import, values='teus', index=['zona', 'bloque', 'estatus'], aggfunc=np.sum, margins=True))
        ornament()
    if len(containers_reefer) > 0:     
        print ('\nREFRIGERADOS')
        print (pd.pivot_table(df2reefer, values='teus', index=['zona', 'bloque', 'estatus', 'frigo'], margins=True, aggfunc=np.sum))
        ornament()  
    if len(containers_ovh) > 0: 
        print ('\nOVH')
        print (pd.pivot_table(df2ovh, values='teus', index=['zona', 'bloque', 'estatus', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_ovhemt) > 0: 
        print ('\nOVHEMT')
        print (pd.pivot_table(df2ovhemt, values='teus', index=['zona', 'bloque', 'estatus', 'tipo'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_mty) > 0:     
        print ('\nVACÍOS')
        print (pd.pivot_table(df2empty, values='teus', index=['zona', 'bloque'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_apto) > 0:     
        print ('\nAPTOS')
        print (pd.pivot_table(df2apto, values='teus', index=['zona', 'bloque'], margins=True, aggfunc=np.sum))
        ornament()
    if len(containers_all) > 0:  
        print ('\nOCUPACIÓN POR LÍNEAS: ')
        print (pd.pivot_table(df2lineasall, values='teus', index=['linea', 'estatus'], margins=True, aggfunc=np.sum))
        ornament()
    
    # Generar archivo EXCEL
    os.chdir(pathtoSave) #Go to the pathfile
    writer = pd.ExcelWriter(f'ocupación_{formato}.xlsx', engine='xlsxwriter')
    workbook = writer.book
    #worksheet = writer.sheets['Ocupación1']
    writer.save()

    with pd.ExcelWriter(f'ocupación_{formato}.xlsx', engine='openpyxl', mode='a') as writer:
        (pd.pivot_table(df2export, values=['teus'], index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=1, startrow=1)
        (pd.pivot_table(df2import, values='teus', index=['zona', 'bloque', 'estatus'], aggfunc=np.sum, margins=True)).to_excel(writer, header=True, startcol=1, startrow=22) #*
        (pd.pivot_table(df2trb, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=6, startrow=1)       
        (pd.pivot_table(df2reefer, values='teus', index=['zona', 'bloque', 'estatus', 'frigo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=6, startrow=18) #* 
        (pd.pivot_table(df2ovh, values='teus', index=['zona', 'bloque', 'estatus','tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=16, startrow=14) #*       
        (pd.pivot_table(df2empty, values='teus', index=['zona', 'bloque'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=12, startrow=1) 
        (pd.pivot_table(df2apto, values='teus', index=['zona', 'bloque'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=6, startrow=30)  #*     
        (pd.pivot_table(df2ovhemt, values='teus', index=['zona', 'bloque', 'estatus', 'tipo'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=16, startrow=1)
        (pd.pivot_table(df2lineasall, values='teus', index=['linea', 'estatus'], margins=True, aggfunc=np.sum)).to_excel(writer, header=True, startcol=22, startrow=1)

    book = openpyxl.load_workbook(f'ocupación_{formato}.xlsx')
    sheet = book.worksheets[1]    

    whitefill = PatternFill(fill_type='solid',
                    start_color='00FFFFFF',
                    end_color='00FFFFFF')    
    for rows in sheet.iter_rows(min_row=1, max_row=60, min_col=1, max_col=30):
        for cell in rows:
            cell.fill = whitefill

    ft = Font(name='Verdana',
                size=12,
                bold = True,
                color='FF000000')                   
    sheet['B1'] = 'EXPORTACIÓN'
    sheet['B1'].font = ft
    sheet['B22'] = 'IMPORTACIÓN'#*
    sheet['B22'].font = ft
    sheet['G1'] = 'TRANSBORDO'
    sheet['G1'].font = ft    
    sheet['G18'] = 'REFRIGERADOS' #*
    sheet['G18'].font = ft
    sheet['Q14'] = 'LLENOS ESPECIALES' #*
    sheet['Q14'].font = ft    
    sheet['M1'] = 'VACÍOS EVACUACIÓN'
    sheet['M1'].font = ft
    sheet['G30'] = 'VACÍOS APTO ALIMENTO' #*
    sheet['G30'].font = ft
    sheet['Q1'] = 'VACÍOS ESPECIALES'
    sheet['Q1'].font = ft
    sheet['W1'] = 'LINEAS'
    sheet['W1'].font = ft
    sheet.title = 'Ocupación'


    #CAPACITY TOTAL TEUS
    cap_llenos = 7150
    cap_vacios = 6460
    cap_total = 13610

    llenos = int(total_llenos)
    pllenos = (llenos / cap_llenos) * 100
    vacios = int(total_vacios)
    pvacios = (vacios / cap_vacios) * 100
    total = llenos + vacios
    ocupacion = (total / cap_total ) * 100

    ft2 = Font(name='Verdana',
                size=11,
                bold = True,
                color='FF000000')
    yellowfill = PatternFill(fill_type='solid',
                    start_color='00FFFF00',
                    end_color='00FFFF00')
    for rows in sheet.iter_rows(min_row=36, max_row=37, min_col=11, max_col=16):
        for cell in rows:
            cell.fill = yellowfill
    for rows in sheet.iter_rows(min_row=39, max_row=39, min_col=11, max_col=16):
        for cell in rows:
            cell.fill = yellowfill        

    sheet['K36'] = f"Teus Conts Llenos {llenos} - ocupación {round(pllenos, 1)}%"
    sheet['K36'].font = ft2
    sheet['K37'] = f"Teus Conts Vacíos {vacios} - ocupación {round(pvacios, 1)}%"
    sheet['K37'].font = ft2
    sheet['K39'] = f"Total Teus Ocupación es {total} - {round(ocupacion, 1)}%"
    sheet['K39'].font = ft


    sheet0 = book.worksheets[0]
    book.remove_sheet(sheet0)
     
    book.save(f'ocupación_{formato}.xlsx')   


    # Generar archivo CSV
    #os.chdir(pathtoSave) #Go to the pathfile
    
    #(pd.pivot_table(df2export, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv')
    #writer = csv.writer(f'Ocupación_{formato}.csv', 'w')
    #writer.writerow(('TRANSBORDOS'))
    #('EXPORTACIÓN').to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2trb, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2import, values='teus', index=['zona', 'bloque', 'estatus'], aggfunc=np.sum, margins=True)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2reefer, values='teus', index=['zona', 'bloque', 'estatus', 'frigo'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2ovh, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2ovhemt, values='teus', index=['zona', 'bloque', 'estatus'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2empty, values='teus', index=['zona', 'bloque'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2apto, values='teus', index=['zona', 'bloque'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')
    #(pd.pivot_table(df2lineasall, values='teus', index=['linea', 'estatus'], margins=True, aggfunc=np.sum)).to_csv(f'Ocupación_{formato}.csv', mode='a')


    #CAPACITY TOTAL TEUS
    #cap_llenos = 7150
    #cap_vacios = 6460
    #cap_total = 13610

    #llenos = int(total_llenos)
    #pllenos = (llenos / cap_llenos) * 100
    #vacios = int(total_vacios)
    #pvacios = (vacios / cap_vacios) * 100
    #total = llenos + vacios
    #ocupacion = (total / cap_total ) * 100

    print (f"\nTotal Llenos {llenos} - ocupacion {round(pllenos, 1)}%")
    print (f"Total Vacíos {vacios} - ocupacion {round(pvacios, 1)}%")
    print (f"\nTotal Ocupación es {total} - {round(ocupacion, 1)}%")

    return (total_vacios, total_llenos)            

   
def ornament():
    print("_" * 50)


def menu1():
    ornament()
    print("OCUPACIÓN EN TEUS")
    ornament()
      

  
def main():  # MAIN
    clear()  
    menu1()
    excel_file= openFile()  # Name of the file to choose
    root.destroy() # Close Dialog
    wb = xlrd.open_workbook(excel_file, encoding_override="cp1252") # Read excel book
    sheet = wb.sheet_by_index(0)
    ornament()  
    ocupacion = ocupacion_list(sheet)
    cap_llenos = 7150
    cap_vacios = 6460
    cap_total = 13610
    llenos = int(ocupacion[1])
    pllenos = (ocupacion[1] / cap_llenos) * 100
    vacios = int(ocupacion[0])
    pvacios = (ocupacion[0] / cap_vacios) * 100
    total = llenos + vacios
    ocupacion = (sum(ocupacion) / cap_total ) * 100
    print (f"\nTotal Llenos {llenos} - ocupacion {round(pllenos, 1)}%")
    print (f"Total Vacíos {vacios} - ocupacion {round(pvacios, 1)}%")
    print (f"\nTotal Ocupación es {total} - {round(ocupacion, 1)}%")


if __name__ == '__main__':
    main()


#os.system("pause") # Press a key to continue 
