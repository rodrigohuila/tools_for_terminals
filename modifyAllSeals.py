#! python3

import os
from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import *  # Prueba borrar cuadro dialogo
from tkinter.ttk import * # Prueba borrar cuadro dialogo
from os.path import basename
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


pathFile = r"S:/OPS/Marítimas/CARPETAS MOTONAVES"
#pathFile = '/home/rodrigo/Downloads/'
col1 = 7  # Column where is sello1
col2 = 8  # Column where is sello2
col3 = 23  # Column where is masselllos
col4 = 0  # Column where is the preestibas
# background color1
style = xlwt.easyxf('pattern: pattern solid, fore_colour yellow')
# background color2
style2 = xlwt.easyxf('pattern: pattern solid, fore_colour pale_blue')
contador = 0 # contador para las preestibas

root = Tk() # Prueba borrar cuadro dialogo
root.update()

def clear():  # ERASE SCREEN
    if os.name == "nt":      # Windows
        os.system("cls")
    else:
        os.system("clear")   # Linux


def getDirectory():  # DIALOG TO CHOOSE A FILE
    sealDirectory = FileDialog.askdirectory(initialdir=pathFile)
    path2 = (sealDirectory)
    # print(path2)
    return(path2)


def openDirectory(sealDirectory):  # DIALOG TO CHOOSE A FILE
    if sealDirectory != "":
        sealFiles = []
        for file in os.listdir(sealDirectory):
            sealFiles.append(file)
    # print(sealFiles)
    return(sealFiles)


def getPreestibas(sealDirectory, path2):  # Get a list of Prrestibas
    for i in range(0, len(sealDirectory)):
        sealFile = (path2 + '/' + sealDirectory[i])
        ext = (os.path.splitext(sealFile)[1])  # get only the extension
        if ext == '.xlsx':
            book = openpyxl.load_workbook(sealFile)
            sheet = book.worksheets[1]  # same sheet in order to can save
            preestibas = []
            for cell in sheet['B']:  # Column B
                preestibas.append(cell.value)
            #print(preestibas)    
            #print('Total de unidades Preestibadas: ' + str(len(preestibas)))
            return(preestibas)


def getAutoFit(sheet):
    ws = sheet
    dims = {}
    for row in ws.rows:
        for cell in row:
            if cell.value:
                dims[cell.column] = max((dims.get(cell.column, 0), len(str(cell.value))))    
    for col, value in dims.items():
        ws.column_dimensions[col].width = value
        


def main():  # MAIN
    global contador
    clear()
    path2 = getDirectory()  # get the path
    sealDirectory = openDirectory(path2)  # Lis of the Directories
    preestibas = getPreestibas(sealDirectory, path2)
    root.destroy() # Prueba de cerrar caja de dialogo

    files = []  # List of files in the directory
    print('No', 'Preestiba')

    for i in range(0, len(sealDirectory)):
        sealFile = (path2 + '/' + sealDirectory[i])
        # print(basename(sealFile))
        #ext = (os.path.splitext(sealFile)[1])  # get extension
        (name, ext) = (os.path.splitext(sealFile))  # get extension

        try:
            # MODIFY TARJA
            if ext == '.xls':
                # print(os.path.splitext(sealFile))
                book = xlrd.open_workbook(sealFile)
                sheet = book.sheet_by_index(1)
                book = copy(book)
                sheet2 = book.get_sheet(1)  # same sheet in order to can save

                                     
                for i in range(sheet.nrows):  # Loop through the sheet
                    # Sello1 without modify
                    massello1 = (repr(sheet.cell_value(i, col1))).replace("'", "")
                    # Sello2 without modify
                    massello2 = (repr(sheet.cell_value(i, col2))).replace("'", "")
                    massello = massello1 + ' ' + massello2

                    sello1 = (repr(sheet.cell_value(i, col1))).replace(' ', '').replace(
                        '-', '').replace("'", "").replace('_', '').strip()
                    sello2 = (repr(sheet.cell_value(i, col2))).replace(' ', '').replace(
                        '-', '').replace("'", "").replace('_', '').strip()
                    selloult = (repr(sheet.cell_value(i, col3))).replace(' ', '').replace(
                        '-', '').replace("'", "").replace('_', '').strip()
                    sello3 = (repr(sheet.cell_value(i, col3))).replace("'", "")

                    sheet2.write(i, col1, sello1)  # Modify all column
                    sheet2.write(i, col2, sello2)  # Modify all column
                    sheet2.write(i, col3, sello3)  # Modify all column

                    if ((sello1[0:2]).isdecimal != True) or ((sello2[0:2]).isdecimal != True):
                        if sello2.find(sello1[0:6]) != -1 and sello1 != '0': # Vacíos nooooooo
                            # Modify cell if is not match
                            sheet2.write(i, col1, '0')
                        elif sello1.find(sello2[0:6]) != -1 and sello2 != '0': # Vacíos nooooooo
                            # Modify cell if is not match
                            sheet2.write(i, col2, '0')

                    if sello1 == sello2:  # Verify repeated seals
                        sheet2.write(i, col1, sello1)  # Modify cell
                        sheet2.write(i, col2, '0')  # Modify cell

                    if sello1 == '' or sello1 == '.':  # Removed '.'
                        sheet2.write(i, col1, '0')  # Modify cell

                    if sello2 == '' or sello2 == '.':  # Removed '.'
                        sheet2.write(i, col2, '0')  # Modify cell

                    if sello1.isdecimal() == False and len(sello1) == 1:
                        sheet2.write(i, col1, '0')  # Modify cell

                    if sello2.isdecimal() == False and len(sello2) == 1:
                        sheet2.write(i, col2, '0')  # Modify cell    
                    

                    # MODIFY MAS SELLOS
                    if ((re.compile(massello)).match(sello3)) != None:
                        # print((re.compile(massello)).match(sello3))
                        if massello1 != '0' or massello2 != '0':
                            sello3 = (sello3.replace(massello, '').replace('.',''))
                            sheet2.write(i, col3, (sello3.lstrip()), style2)  # Modify cell
                            #print(sello3)
                    elif ((re.compile(massello1)).match(sello3)) != None:
                        # print((re.compile(massello1)).match(sello3))
                        if massello1 != '0' and massello1 != '.':
                            sello3 = (sello3.replace(massello1, ''))
                            sheet2.write(i, col3, (sello3.lstrip()), style2)  # Modify cell
                            #print(sello3)
                    elif ((re.compile(massello2)).match(sello3)) != None:
                        # print((re.compile(massello2)).match(sello3))
                        if massello2 != '0' and massello2 != '.':
                            sello3 = (sello3.replace(massello2, ''))
                            sheet2.write(i, col3, (sello3.lstrip()), style)  # Modify cell
                            #print(sello3)

                    if sello3.isdecimal() == False and len(sello2) == 1:
                        sheet2.write(i, col3, '0')  # Modify cell        

                    if sello3 == ' ' or sello3 == '' or  sello3 == '.':
                        sheet2.write(i, col3, '0')  # Modify cell



                    # COLORED PREESTIBAS
                    if type(preestibas) == list:
                        contenedor = (
                            ((repr(sheet.cell_value(i, col4)))).replace("'", ""))
                        if contenedor in preestibas:
                            contador = contador + 1 
                            print(contador, contenedor)
                            sheet2.write(i, col4, contenedor, style)  # Modify cell
                            #sheet2.write(i, col1, finalSeal1, style)  # Modify cell
                            #sheet2.write(i, col2, finalSeal2, style)  # Modify cell
                            #sheet2.write(i, col3, finalSeal3, style)  # Modify cell
            

                #getAutoFit(sheet2)
                files.append(sealFile)
                book.save(sealFile)  # Save

                #newname = (name + 'copy' + ext) # Save the file with other name
                #files.append(newname)
                #book.save(newname)  # Save


        #Excepción de columna mas sellos desaparecida
        except:
                for i in range(sheet.nrows):  # Loop through the sheet
                    # Sello1 without modify
                    massello1 = (repr(sheet.cell_value(i, col1))).replace("'", "")
                    # Sello2 without modify
                    massello2 = (repr(sheet.cell_value(i, col2))).replace("'", "")
                    massello = massello1 + ' ' + massello2

                    sello1 = (repr(sheet.cell_value(i, col1))).replace(' ', '').replace(
                        '-', '').replace("'", "").replace('_', '').strip()
                    sello2 = (repr(sheet.cell_value(i, col2))).replace(' ', '').replace(
                        '-', '').replace("'", "").replace('_', '').strip()
                    
                    sheet2.write(i, col1, sello1)  # Modify all column
                    sheet2.write(i, col2, sello2)  # Modify all column
                    
                    if ((sello1[0:2]).isdecimal != True) or ((sello2[0:2]).isdecimal != True):
                        if sello2.find(sello1[0:6]) != -1 and sello1 != '0': # Vacíos nooooooo
                            # Modify cell if is not match
                            sheet2.write(i, col1, '0')
                        elif sello1.find(sello2[0:6]) != -1 and sello2 != '0': # Vacíos nooooooo
                            # Modify cell if is not match
                            sheet2.write(i, col2, '0')

                    if sello1 == sello2:  # Verify repeated seals
                        sheet2.write(i, col1, sello1)  # Modify cell
                        sheet2.write(i, col2, '0')  # Modify cell

                    if sello1 == '' or sello1 == '.':  # Removed '.'
                        sheet2.write(i, col1, '0')  # Modify cell

                    if sello2 == '' or sello2 == '.':  # Removed '.'
                        sheet2.write(i, col2, '0')  # Modify cell
                        
                    if sello1.isdecimal() == False and len(sello1) == 1:
                        sheet2.write(i, col1, '0')  # Modify cell

                    if sello2.isdecimal() == False and len(sello2) == 1:
                        sheet2.write(i, col2, '0')  # Modify cell        

                    
                    # COLORED PREESTIBAS
                    if type(preestibas) == list:
                        contenedor = (
                            ((repr(sheet.cell_value(i, col4)))).replace("'", ""))    
                        if contenedor in preestibas:
                            contador = contador + 1 
                            print(contador, contenedor)
                            sheet2.write(i, col4, contenedor, style)  # Modify cell
                        
          
                getAutoFit(sheet2)
                files.append(sealFile)
                book.save(sealFile)  # Save

           
    #Print Results
    print('\nTotal de unidades en la Plantilla de Preestibas: ' + str(contador))
    print('\nArchivos procesados: ')

    for i in range(len (files)):
        print(basename(files[i]))

        

if __name__ == '__main__':
    main()


os.system("pause") # Press a key to continue 
    
