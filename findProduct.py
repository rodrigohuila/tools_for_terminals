#! python3

import os
from getpass import getuser
import openpyxl
import xlrd
from os.path import basename
from xlutils.copy import copy
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.styles import PatternFill
import time
from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import Tk  # Prueba borrar cuadro dialogo
import re


#path = rf'/home/{getuser()}/Documents/curso_python/MyScripts/controloperacionesproyect'
userName = getuser()
startpath = rf'S:\OPS\Marítimas\Control Operaciones/'
root = Tk()  # Borrar cuadro dialogo
root.update()


def clear():  # ERASE SCREEN
    if os.name == "nt":      # Windows
        os.system("cls")
    else:
        os.system("clear")   # Linux


def openDirectory(pathFiles):  # DIALOG TO CHOOSE A FILE
    if pathFiles != "":
        excelFiles = []
        for file in os.listdir(pathFiles):
            excelFiles.append(file)
    # print(excelFiles)    
    return(excelFiles)


def getFile(startpath, nick):  # DIALOG TO CHOOSE A FILE
    if nick == 'excel':
        # file = FileDialog.askopenfilename(
        #    initialdir=path,
        #    filetypes=(('xlsx', '*.xlsx'), ('all files', '*.*')),
        #    title='Escoja el archivo a procesar')
        filespath = FileDialog.askdirectory(
            initialdir=startpath,
            title='Escoja la carpeta donde se encuentran el/los archivo(s) a procesar'
        )
        files = openDirectory(filespath)
        #print (f'el path del directorio es: {filespath}')
        return(files, filespath)
        

    elif nick == 'base':
        file = FileDialog.askopenfilename(
            initialdir=startpath,
            filetypes=(('xls', '*.xls'), ('all files', '*.*')),
            title='Escoja la base de datos para confrontar la información'
        )
        # print(f'\nBase de Datos escogida: {basename(file)}')
        path = os.path.dirname(file)
        #print (f'el path de la base de datos: {path}')
        return(file, path)

        # except:
        #    print('No has escogido ningún archivo, por favor despierta')


def getproducts(wb, sheet1, sheet2=None):
    '''From a excel extension xls'''
    products = []
    omite = re.compile(
        r'\d+\b(AND|Y|WITH|CON|OF|DE|THE|A|AL|LAS|LOS|LA|LO|EL|YOUR\'S|SU|CONSOLIDATE CARGO|CARGA CONSOLIDADA|CONSOLIDATE|CONSOLIDADA|PC|PCS|AUTOMATIC|AUTOMATICO|AUTOMATICS|AUTOMATICOS|&|\*|-|_)\b', re.IGNORECASE)
    for i in range(sheet1.nrows):
        products.append(omite.sub('', sheet1.cell_value(i, 3).upper().strip()))
        products.append(omite.sub('', sheet1.cell_value(i, 4).upper().strip()))
    for i in range(sheet2.nrows):
        products.append(omite.sub('', sheet2.cell_value(i, 3).upper().strip()))
    totalproducts = len(products) - 3
    print(f'\nCantidad de productos en la base de datos es: {totalproducts}')
    # print(products)
    return(products)


def checkproducts(file, basedata, wb, sheet, wordposition=0):
    matches = []
    nomatches = []
    omite = re.compile(
        r'\b(AND|Y|WITH|CON|IN|EN|OF|DE|THE|AL|LAS|LOS|LA|LO|EL|YOUR\'S|SU|SACO|BAG|SACOS|BAGS|CARGA|CARGO|VACIO|EMPTY|MTY)\b')  # Only if the complete word
    for cell in sheet['I']:
        if cell.value != None:
            celda = cell.value.upper().strip()
            # listwords = cell.value.split()
            listwords = re.split(r'[\W]+', omite.sub('', celda))
            # print(listwords)
            for palabra in listwords:
                if palabra != '':
                    if any(palabra in word for word in basedata):
                        #exact = re.compile(r'\b(' + palabra + r')\b')
                        # for word in basedata:
                        # if re.match(exact, word) != None:
                        matches.append(celda)
                        if celda in nomatches:
                            cell.font = Font(color='00000000',
                                             italic=False, bold=False)
                            nomatches.remove(celda)
                    else:
                        if celda not in matches:
                            if celda not in nomatches:
                                nomatches.append(celda)
                            cell.font = Font(color='00FF0000',
                                             italic=True, bold=True)
                            cell.fill = PatternFill(
                                fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')
        else:
            cell.fill = PatternFill(
                fill_type='solid', start_color='00FFFF00', end_color='00FFFF00')

    # wb.save(filename='checkedfile.xls')
    wb.save(file)
    print(f'\ntotal de inconsistencias: {len(set(nomatches))}')
    for item in enumerate(set(nomatches), start=1):
        print(item)


def readFile2(files, products, path, sheetNumber=0):
    '''Read excel files with ext xlsx'''
    for i in range(0, len(files)):
        file = (path + '/' + files[i])
        ext = (os.path.splitext(file)[1])  # get only the extension
        if ext == '.xlsx':
            wb = openpyxl.load_workbook(file)
            sheet = wb.worksheets[0]
            print(
                f'\nNúmeros de búsquedas de productos realizadas: {sheet.max_row - 9}\nen el archivo: {os.path.basename(file)}')
            checkproducts(file, products, wb, sheet)
    # return (wb, sheet)


def readFile(file):
    '''Read excel files with ext xls'''
    wb = xlrd.open_workbook(
        file, encoding_override="cp1252")  # Read excel book
    sheet1 = wb.sheet_by_index(0)
    sheet2 = wb.sheet_by_index(2)
    basename = os.path.basename(file)
    print(f'El archivo de la base de datos es: {basename}')
    return (wb, sheet1, sheet2)


def ornament():
    print("_" * 60)


def menu1():
    ornament()
    print ()
    print (' CHECKING PRODUCTS '.center(76, '/'))
    ornament()

def greetings():
    user = userName.upper()
    userUpper = re.sub('([A-Z])', r' \1', user)
    print ()
    print (f'H E L L O {userUpper}'.center(80))
    print ('No olvidar colocar los productos a revisar en la columna: I'.center(80))
    ornament()


def main():
    clear()
    ornament()
    menu1()
    greetings()
    baseDatos = getFile(startpath, 'base')
    excel_files = getFile(startpath, 'excel')
    root.destroy()  # cerrar caja de dialogo
    '''xls'''
    book = readFile(baseDatos[0])  # get the wb and the sheet1 and the sheet2
    products = getproducts(book[0], book[1], book[2])
    start_time = time.time()  # Star Proces of Checking
    '''xlsx'''
    readFile2(excel_files[0], products, excel_files[1])
    # checkproducts(products, file2[0], file2[1])
    print(
        f'\nTiempo en que realizó la búsqueda: {round(time.time() - start_time,2)}--- %s segundos --- %')


if __name__ == '__main__':
    main()

os.system('pause')  # Press a key to continue
