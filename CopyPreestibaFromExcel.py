#! python3

from tkinter import filedialog as FileDialog
from tkinter import messagebox as MessageBox
from tkinter import *  # Prueba borrar cuadro dialogo
from tkinter.ttk import * # Prueba borrar cuadro dialogo
from os.path import basename
import pyautogui
import pyperclip
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
import re
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import os
import time
from getpass import getuser


userName = getuser()
path = r"S:\OPS\Dispatchers\Preestibas"
#file = 'PREESTIBAS FISICAS.xlsx'
#style = xlwt.easyxf('pattern: pattern solid, fore_colour yellow') # background color
root = Tk() # Borrar cuadro dialogo
root.update()
i = 1
espacios = 80


def clear():  # ERASE SCREEN
    if os.name == "nt":      # Windows
        os.system("cls")
    else:
        os.system("clear")   # Linux
        

def getFile(path):  # DIALOG TO CHOOSE A FILE
    file = FileDialog.askopenfilename(
        initialdir=path,
        filetypes=(('xlsx', '*.xlsx'), ('all files', '*.*')),
        title='E s c o j a__e l__a r c h i v o__d e__P r e e s t i b a s__a__ p r o c e s a r'
     )
    #print('\nEl nombre del archivo inicial es: ' + basename(file) + '\n')
    print (f'Archivo excel selecionado: {file}')
    return(file)


def getPreestibas(file):  # Get a list of Prrestibas
    book = openpyxl.load_workbook(file)
    sheet = book.worksheets[1]  
    preestibas = []
    for cell in sheet['A']:  # Column A
        preestibas.append(cell.value)
    return(preestibas)


def writePreestibas(opc, contenedor, i):
    while True:
        try:
            if opc == 1:
                pyautogui.click(702, 384) # clic contenedor Monitor 1
                pyautogui.typewrite(contenedor[:-1])
                #time.sleep(1)
                pyautogui.click(1081, 606) # clic aceptar Monitor 1
                i += 1
                if i < 10:
                    print(f'  {i}    {contenedor}')
                elif i >= 10:
                    print(f' {i}    {contenedor}')
                #pyautogui.click(1168, 606) # clic Cancelar Monitor 1
                break
            elif opc == 2:
                pyautogui.click(2622, 375) # clic contenedor Monitor 2
                pyautogui.typewrite(contenedor[:-1])
                #time.sleep(1)
                pyautogui.click(3004, 606) # clic aceptar Monitor 2
                i += 1
                if i < 10:
                    print(f'  {i}    {contenedor}')
                elif i >= 10:
                    print(f' {i}    {contenedor}')
                #pyautogui.click(3095, 606) # clic Cancelar Monitor 2
                break
        except ValueError:
            ornament()
            print ('Opción invalida')
            clear()
            main()

def ornament():
    print("_" * espacios)


def menu1():
    ornament()
    print ()
    print(" TYPING PREESTIBAS FROM EXCEL ".center(espacios, '/'))
    ornament()


def greetings():
    user = userName.upper()
    userUpper = re.sub('([A-Z])', r' \1', user)
    print ()
    print (f'H E L L O O O O      {userUpper}'.center(espacios))
    print ('No olvidar crear un Nuevo Posicionado con Tipo y Movimiento antes de iniciar')
    print ('y de copiar *solo* contenedores en la hoja 2 del excel del archivo de Preestibas')
    print ('ADVERTENCIA: Si existe un posicionado de la lista no se digitaran más unidades')
    ornament()



def main():  # MAIN
    clear()
    menu1()
    file = getFile(path)
    root.destroy() # Prueba de cerrar caja de dialogo
    preestibas = getPreestibas(file)
    book = xlrd.open_workbook(file)
    sheet = book.sheet_by_index(1)
    book = copy(book)
    sheet2 = book.get_sheet(1)  # same sheet in order to can save

    greetings()

    opc = input ('\nEscoja Monitor No 1 o Monitor No 2: ')

    ornament()
    print ('Cant   Contenedor')
    ornament()

    for i in range(sheet.nrows):  # Loop through the sheet
        contenedor = (repr(sheet.cell_value(i, 0)).replace("'", ""))
        if type(preestibas) == list:
            if contenedor in preestibas:
                writePreestibas(int(opc), contenedor, i)
                
    print('\nTotal preestibas digitadas:' + str(len(preestibas)))           

    if int(opc) == 1:
        pyautogui.click(1168, 606) # clic Cancelar Monitor 1
    elif int(opc) == 2 :
        pyautogui.click(3095, 606) # clic Cancelar Monitor 2
        


if __name__ == '__main__':
    main()


os.system("pause") # Press a key to continue 
